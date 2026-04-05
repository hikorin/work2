import sys
import openpyxl

file_path = r'C:\Users\hikoi\Downloads\かえし原価計算.xlsx'
output_path = r'C:\Data\repo\work2\docs\kaeshi_logic.md'

try:
    with open(output_path, 'w', encoding='utf-8') as f:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        
        f.write('# かえし原価計算 - ロジック解析結果\n\n')
        for sheet_name in wb.sheetnames:
            f.write(f'## シート: {sheet_name}\n')
            ws_f = wb[sheet_name]
            ws_d = wb_data[sheet_name]
            for row_f, row_d in zip(ws_f.iter_rows(), ws_d.iter_rows()):
                for cf, cd in zip(row_f, row_d):
                    if cf.value is not None:
                        if str(cf.value).startswith('='):
                            f.write(f'- **{cf.coordinate}**: `{cf.value}` (評価結果: {cd.value})\n')
                        else:
                            f.write(f'- **{cf.coordinate}**: {cf.value}\n')
            f.write('\n')
except Exception as e:
    print(f'Error: {e}')
