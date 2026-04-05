# coding: utf-8
import sys, subprocess
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-q', 'openpyxl'])
    import openpyxl

file_path = r'C:\Users\hikoi\Downloads\かえし原価計算.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    # data_only=True evaluates the formulas to string values (if the file was saved by Excel with cached values)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        print(f'\n=== Sheet: {sheet_name} ===')
        ws_f = wb[sheet_name]
        ws_d = wb_data[sheet_name]
        for row_f, row_d in zip(ws_f.iter_rows(), ws_d.iter_rows()):
            for cf, cd in zip(row_f, row_d):
                if cf.value is not None:
                    # Check if it's a formula
                    if str(cf.value).startswith('='):
                        print(f'{cf.coordinate}: {cf.value}  (結果: {cd.value})')
                    else:
                        print(f'{cf.coordinate}: {cf.value}')
except Exception as e:
    print(f'Error reading Excel file: {e}')
